import tkinter as tk
from tkinter import filedialog, messagebox

import sqlite3
import pandas as pd
from openpyxl import load_workbook
import threading

def on_click_output(root, name_db):
    
    def write_files(root, name_db, file_path, loading):

        conn = sqlite3.connect(name_db)
        cursor = conn.cursor()

        vacancy_sheet = "Vacancy"
        person_sheet = "Person"

        wb = load_workbook(file_path)

        try:
            wb.remove(wb[person_sheet])
            wb.save(file_path)
        except Exception:
            pass

        try:
            wb.remove(wb[vacancy_sheet])
            wb.save(file_path)
        except Exception:
            pass

        wb.create_sheet(title=person_sheet)
        wb.create_sheet(title=vacancy_sheet)

        try:
            wb.save(file_path)
        except Exception:
            root.after(0, lambda: messagebox.showerror("Помилка", f'Файл: "{file_path}" вже відкритий!'))
            conn.close()
            root.after(0, loading.destroy)
            return

        cursor.execute('''
                SELECT

                code_year."назва_кафедри",
                code_year."роки",

                person."ім_я",
                person."прізвище",
                person."по-батькові",

                job."посада",
                job."вчене_звання_вчена_ступінь",
                job."знак_сумісності",
   
                first_semester."ставка",
                first_semester."лекції",
                first_semester."практичні_(семінарські)_заняття",
                first_semester."лабораторні_роботи",
                first_semester."екзамени",
                first_semester."консультації_перед_екзаменами",
                first_semester."заліки",
                first_semester."кваліфікаційна_робота",
                first_semester."атестаційний_екзамен",
                first_semester."виробнича_практика",
                first_semester."навчальна_практика",
                first_semester."поточні_консультації",
                first_semester."індивідуальні",
                first_semester."курсові_роботи",
                first_semester."аспірантські_екзамени",
                first_semester."керівництво_аспірантами",
                first_semester."консультування_докторантів_здобувачів",
                first_semester."керівництво_ФПК",
                first_semester."робота_приймальної_комісії",
                first_semester."інше",
                first_semester."всього",

                second_semester."ставка",
                second_semester."лекції",
                second_semester."практичні_(семінарські)_заняття",
                second_semester."лабораторні_роботи",
                second_semester."екзамени",
                second_semester."консультації_перед_екзаменами",
                second_semester."заліки",
                second_semester."кваліфікаційна_робота",
                second_semester."атестаційний_екзамен",
                second_semester."виробнича_практика",
                second_semester."навчальна_практика",
                second_semester."поточні_консультації",
                second_semester."індивідуальні",
                second_semester."курсові_роботи",
                second_semester."аспірантські_екзамени",
                second_semester."керівництво_аспірантами",
                second_semester."консультування_докторантів_здобувачів",
                second_semester."керівництво_ФПК",
                second_semester."робота_приймальної_комісії",
                second_semester."інше",
                second_semester."всього",

                academic_year."ставка",
                academic_year."лекції",
                academic_year."практичні_(семінарські)_заняття",
                academic_year."лабораторні_роботи",
                academic_year."екзамени",
                academic_year."консультації_перед_екзаменами",
                academic_year."заліки",
                academic_year."кваліфікаційна_робота",
                academic_year."атестаційний_екзамен",
                academic_year."виробнича_практика",
                academic_year."навчальна_практика",
                academic_year."поточні_консультації",
                academic_year."індивідуальні",
                academic_year."курсові_роботи",
                academic_year."аспірантські_екзамени",
                academic_year."керівництво_аспірантами",
                academic_year."консультування_докторантів_здобувачів",
                academic_year."керівництво_ФПК",
                academic_year."робота_приймальної_комісії",
                academic_year."інше",
                academic_year."всього",
                academic_year."розподіл_ставок_навчального_навантаження",
                job."НПП_ПП"
 
                FROM code_year

                JOIN job ON code_year.id = job."код_рік_id" 
                JOIN person ON job.id = person."job_id" 
                JOIN first_semester ON job.id = first_semester."job_id" 
                JOIN second_semester ON job.id = second_semester."job_id" 
                JOIN academic_year ON job.id = academic_year."job_id" 

                ORDER BY code_year."назва_кафедри";

            ''')

        persons = [list(person) for person in cursor.fetchall()]

        cursor.execute('''
            SELECT

            code_year."назва_кафедри",
            code_year."роки",

            vacancy."назва",
            vacancy."номер",

            job."посада",
            job."вчене_звання_вчена_ступінь",
            job."знак_сумісності",
   
            first_semester."ставка",
            first_semester."лекції",
            first_semester."практичні_(семінарські)_заняття",
            first_semester."лабораторні_роботи",
            first_semester."екзамени",
            first_semester."консультації_перед_екзаменами",
            first_semester."заліки",
            first_semester."кваліфікаційна_робота",
            first_semester."атестаційний_екзамен",
            first_semester."виробнича_практика",
            first_semester."навчальна_практика",
            first_semester."поточні_консультації",
            first_semester."індивідуальні",
            first_semester."курсові_роботи",
            first_semester."аспірантські_екзамени",
            first_semester."керівництво_аспірантами",
            first_semester."консультування_докторантів_здобувачів",
            first_semester."керівництво_ФПК",
            first_semester."робота_приймальної_комісії",
            first_semester."інше",
            first_semester."всього",

            second_semester."ставка",
            second_semester."лекції",
            second_semester."практичні_(семінарські)_заняття",
            second_semester."лабораторні_роботи",
            second_semester."екзамени",
            second_semester."консультації_перед_екзаменами",
            second_semester."заліки",
            second_semester."кваліфікаційна_робота",
            second_semester."атестаційний_екзамен",
            second_semester."виробнича_практика",
            second_semester."навчальна_практика",
            second_semester."поточні_консультації",
            second_semester."індивідуальні",
            second_semester."курсові_роботи",
            second_semester."аспірантські_екзамени",
            second_semester."керівництво_аспірантами",
            second_semester."консультування_докторантів_здобувачів",
            second_semester."керівництво_ФПК",
            second_semester."робота_приймальної_комісії",
            second_semester."інше",
            second_semester."всього",

            academic_year."ставка",
            academic_year."лекції",
            academic_year."практичні_(семінарські)_заняття",
            academic_year."лабораторні_роботи",
            academic_year."екзамени",
            academic_year."консультації_перед_екзаменами",
            academic_year."заліки",
            academic_year."кваліфікаційна_робота",
            academic_year."атестаційний_екзамен",
            academic_year."виробнича_практика",
            academic_year."навчальна_практика",
            academic_year."поточні_консультації",
            academic_year."індивідуальні",
            academic_year."курсові_роботи",
            academic_year."аспірантські_екзамени",
            academic_year."керівництво_аспірантами",
            academic_year."консультування_докторантів_здобувачів",
            academic_year."керівництво_ФПК",
            academic_year."робота_приймальної_комісії",
            academic_year."інше",
            academic_year."всього",
            academic_year."розподіл_ставок_навчального_навантаження",
            job."НПП_ПП"
 
            FROM code_year

            JOIN job ON code_year.id = job."код_рік_id" 
            JOIN vacancy ON job.id = vacancy."job_id" 
            JOIN first_semester ON job.id = first_semester."job_id" 
            JOIN second_semester ON job.id = second_semester."job_id" 
            JOIN academic_year ON job.id = academic_year."job_id" 

            
            ORDER BY code_year."назва_кафедри";

            ''')

        vacancies = [list(vacancy) for vacancy in cursor.fetchall()]

        titles = ["Назва кафедри","Роки","Ім'я","Прізвище","По-батькові","Посада","Вчене звання, ступінь", "Знак сумісності",
        "Ставка (перший семестр)" , "Лекції (перший семестр)", "Практичні (семінарські) заняття (перший семестр)",
        "Лабораторні роботи (перший семестр)", "Eкзамени (перший семестр)", "Kонсультації перед екзаменами (перший семестр)",
        "Заліки (перший семестр)", "Кваліфікаційні роботи (перший семестр)", "Атестаційні екзамени (перший семестр)",
        "Виробнича практика (перший семестр)", "Навчальна практика (перший семестр)", "Поточні консультації (перший семестр)",
        "Індивідуальне (перший семестр)", "Курсові роботи (перший семестр)", "Аспірантські екзамени (перший семестр)", 
        "Керівництво аспірантами (перший семестр)", "Консультування докторантів, здобувачів (перший семестр)", "Kерівництво ФПК (перший семестр)",
        "Робота приймальної комісії (перший семестр)", "Інше (перший семестр)", "Всього (перший семестр)",
        
        "Ставка (другий семестр)" , "Лекції (другий семестр)", "Практичні (семінарські) заняття (другий семестр)",
        "Лабораторні роботи (другий семестр)", "Eкзамени (другий семестр)", "Kонсультації перед екзаменами (другий семестр)",
        "Заліки (другий семестр)", "Кваліфікаційні роботи (другий семестр)", "Атестаційні екзамени (другий семестр)",
        "Виробнича практика (другий семестр)", "Навчальна практика (другий семестр)", "Поточні консультації (другий семестр)",
        "Індивідуальне (другий семестр)", "Курсові роботи (другий семестр)", "Аспірантські екзамени (другий семестр)", 
        "Керівництво аспірантами (другий семестр)", "Консультування докторантів, здобувачів (другий семестр)", "Kерівництво ФПК (другий семестр)",
        "Робота приймальної комісії (другий семестр)", "Інше (другий семестр)", "Всього (другий семестр)",
        
        "Ставка (академічний рік)" , "Лекції (академічний рік)", "Практичні (семінарські) заняття (академічний рік)",
        "Лабораторні роботи (академічний рік)", "Eкзамени (академічний рік)", "Kонсультації перед екзаменами (академічний рік)",
        "Заліки (академічний рік)", "Кваліфікаційні роботи (академічний рік)", "Атестаційні екзамени (академічний рік)",
        "Виробнича практика (академічний рік)", "Навчальна практика (академічний рік)", "Поточні консультації (академічний рік)",
        "Індивідуальне (академічний рік)", "Курсові роботи (академічний рік)", "Аспірантські екзамени (академічний рік)", 
        "Керівництво аспірантами (академічний рік)", "Консультування докторантів, здобувачів (академічний рік)", "Kерівництво ФПК (академічний рік)",
        "Робота приймальної комісії (академічний рік)", "Інше (академічний рік)", "Всього (академічний рік)", "Розподіл ставок навчального навантаження", "НПП_ПП"]

        for i,title in enumerate(titles,1):
            wb[person_sheet].cell(row=1, column=i, value=title)

        
        titles = titles[5:]

        titles = ["Назва кафедри","Роки","Назва вакансії","Номер вакансії"] + titles

        for i,title in enumerate(titles,1):
            wb[vacancy_sheet].cell(row=1, column=i, value=title)

        
        for i, person in enumerate(persons, 2):
            for j, temp in enumerate(person, 1):
                wb[person_sheet].cell(row=i, column=j, value=temp)

        for i, vacancy in enumerate(vacancies, 2):
            for j, temp in enumerate(vacancy, 1):
                wb[vacancy_sheet].cell(row=i, column=j, value=temp)

        wb.save(file_path)
               
        conn.close()
        root.after(0, loading.destroy)

    from src.settings.settings import loading_window
    from src.menus.main import main_menu

    file_path = filedialog.askopenfilename(
        title="Оберіть Excel файл",
        filetypes=[("Excel File", "*.xlsx")]
    )
    
    if not file_path:
        messagebox.showinfo("Помилка", "Файл не обрано або має валідну помилку в назві!")
        return

        

    loading = loading_window(root)
    threading.Thread(target=write_files, args=(root, name_db, file_path, loading), daemon=True).start()
   
    return