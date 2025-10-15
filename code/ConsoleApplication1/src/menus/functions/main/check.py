import tkinter as tk
from tkinter import filedialog, messagebox

import sqlite3
import pandas as pd
from openpyxl import load_workbook
import threading

def on_click_check(root, name_db):
    from src.settings.settings import loading_window

    class Person:

        def __init__(self, surname, name, patronomic, position, department, sign="", bid=0, total_year=0, job_id=0, npp_pp="ПП"):

            self.__surname = str(surname).replace(' ','')
            self.__name = str(name).replace(' ','')
            self.__patronomic = str(patronomic).replace(' ','')

            position = str(position).lower().replace(' ','')

            if position: 
                if position[0] == 'з':
                    self.__position = "заф. кафедри"
                elif position[0].lower() == 'п':
                    self.__position = "професор"
                elif position[0].lower() == 'д':
                    self.__position = "доцент"
                elif position[0].lower() == 'с':
                    self.__position = "ст. викладач"
                elif position[0].lower() == 'а':
                    self.__position = "асистент"
                elif position[0].lower() == 'в':
                    if position[1].lower()  == '.':
                        self.__position = "в.о. заф. кафедри"
                    elif position[1].lower() == 'и':
                        self.__position = "викладач"
                    else:
                       self.__position = ""
                else:
                    self.__position = ""
            else:
                self.__position = ""

            self.__department = str(department).replace(' ','')

            self.__sign = str(sign).replace(' ','')
            if self.__sign:
                self.__sign = "сум."

            self.__bid = float(str(bid).replace(' ','').replace(',','.'))

            self.__total_year = float(str(total_year).replace(' ','').replace(',','.'))

            self.__job_id = int(str(job_id).replace(' ',''))

            self.__less_time = float(0.0)

            self.__npp_pp = "НПП" if str(npp_pp).lower().replace(' ','')[0] == "н" else "ПП"


        def set_npp_pp(self,value):
            self.__npp_pp = "НПП" if str(value).lower().replace(' ','')[0] == "н" else "ПП"

        def get_npp_pp(self):
            return str(self.__npp_pp).replace(' ','')

        def get_surname(self):
            return str(self.__surname).replace(' ','')

        def get_name(self):
            return str(self.__name).replace(' ','')

        def get_patronomic(self):
            return str(self.__patronomic).replace(' ','')

        def get_position(self):
            return str(self.__position).replace(' ','')

        def get_department(self):
            return str(self.__department).replace(' ','')

        def get_bid(self):
            return float(str(self.__bid).replace(' ','').replace(',','.'))

        def get_total_year(self):
            return float(str(self.__total_year).replace(' ','').replace(',','.'))

        def get_job_id(self):
            return int(str(self.__job_id).replace(' ',''))

        def get_less_time(self):
            return float(str(self.__less_time).replace(' ','').replace(',','.'))

        def get_sign(self):
            return f"{self.__sign}"
            

        def __str__(self):
            return f"name = {self.get_name()};\nsurname = {self.get_surname()};\npatronomic = {self.get_patronomic()};\nless_time = {self.get_less_time()};\nposition = {self.__position};\ndepartment = {self.__department};\nbid = {self.__bid};\ntotal year = {self.__total_year}.\nsign = {self.__sign}\n\n"
            
       


        def add_less_time(self, value):
            self.__less_time += float(str(value).replace(' ','').replace(',','.'))

      
        def __eq__(self, other):
            if not isinstance(other, Person):
                return NotImplemented
           
            if self.__job_id!=0 and other.__job_id!=0: 
                return self.__job_id == other.__job_id

            return (self.__name == other.__name and self.__surname == other.__surname and self.__patronomic == other.__patronomic
            and  self.__position == other.__position and  self.__department == other.__department and  self.__sign == other.__sign)

    class Vacancy:
        def __init__(self, number, position, department, sign="", bid=0, total_year=0, job_id=0,npp_pp="ПП"):

            self.__number = int(str(number).replace(' ',''))

            position = str(position).lower().replace(' ','')

            if position: 
                if position[0] == 'з':
                    self.__position = "заф. кафедри"
                elif position[0].lower() == 'п':
                    self.__position = "професор"
                elif position[0].lower() == 'д':
                    self.__position = "доцент"
                elif position[0].lower() == 'с':
                    self.__position = "ст. викладач"
                elif position[0].lower() == 'а':
                    self.__position = "асистент"
                elif position[0].lower() == 'в':
                    if position[1].lower()  == '.':
                        self.__position = "в.о. заф. кафедри"
                    elif position[1].lower() == 'и':
                        self.__position = "викладач"
                    else:
                       self.__position = ""
                else:
                    self.__position = ""
            else:
                self.__position = ""

            self.__department = str(department).replace(' ','')

            self.__sign = str(sign).replace(' ','')
            if self.__sign:
                self.__sign = "сум."

            self.__bid = float(str(bid).replace(' ','').replace(',','.'))

            self.__total_year = float(str(total_year).replace(' ','').replace(',','.'))

            self.__job_id = int(str(job_id).replace(' ',''))

            self.__npp_pp = "НПП" if str(npp_pp).lower().replace(' ','')[0] == "н" else "ПП"

            self.__less_time = float(0.0)


        def set_npp_pp(self,value):
            self.__npp_pp = "НПП" if str(value).lower().replace(' ','')[0] == "н" else "ПП"

        def get_npp_pp(self):
            return str(self.__npp_pp).replace(' ','')


        def get_number(self):
            return int(str(self.__number).replace(' ',''))

        def get_position(self):
            return str(self.__position).replace(' ','')

        def get_department(self):
            return str(self.__department).replace(' ','')

        def get_bid(self):
            return float(str(self.__bid).replace(' ',''))

        def get_total_year(self):
            return float(str(self.__total_year).replace(' ','').replace(',','.'))

        def get_job_id(self):
            return int(str(self.__job_id).replace(' ',''))

        def get_less_time(self):
            return float(str(self.__less_time).replace(' ','').replace(',','.'))  

        def __str__(self):
            return f"name = Вакансія;\nnumber = {self.get_number()};\nless_time = {self.get_less_time()};\nposition = {self.__position};\ndepartment = {self.__department};\nbid = {self.__bid};\ntotal year = {self.__total_year}.\nsign = {self.__sign}\n\n"
           

        def add_less_time(self, value):
            self.__less_time += float(str(value).replace(' ','').replace(',','.'))

        def __eq__(self, other):
            if not isinstance(other, Vacancy):
                return NotImplemented
           
            if self.__job_id!=0 and other.__job_id!=0: 
                return self.__job_id == other.__job_id

            return (self.__number== other.__number
            and  self.__position == other.__position and  self.__department == other.__department and  self.__sign == other.__sign)

        def get_sign(self):
            return f"{self.__sign}"

    class Load:
        def __init__(self, name, npp_pp, value):

            self.__name = str(name)
            self.__npp_pp = str(npp_pp).replace(' ','')
            self.__value = float(str(value).replace(' ','').replace(',','.'))

        def get_name(self):
            return str(self.__name)

        def get_npp_pp(self):
            return str(self.__npp_pp).replace(' ','')

        def get_value(self):
            return float(str(self.__value).replace(' ','').replace(',','.'))

        def __str__(self):
            return f"name  = {self.__name};\tnpp_pp = {self.__npp_pp};\tvalue = {self.__value}.\n"

    def check_file(root, name_db, file_path):
        loading = loading_window(root)
        conn = sqlite3.connect(name_db)
        cursor = conn.cursor()

        def get_sheets(cursor):
            cursor.execute("""
                SELECT DISTINCT name
                FROM load
                WHERE name != 'Максимум' AND  name != 'Мінімум' AND name != 'Середнє навантаження' AND name != 'Максимум перевищення у %' AND name != 'Рівномірність розподілу' AND name != 'Керівники аспірантів'
                ORDER BY name
            """)
            result = cursor.fetchall()
            sheets = []
            for value in result:
                sheets.append(str(value[0]))
            return sheets
                
        def get_cell_str(wb, sheet_name,row,col):
            return str(wb[sheet_name].cell(row=row, column=col).value) if str(wb[sheet_name].cell(row=row, column=col).value).replace(' ','').lower() != 'none' else ""
            
        def get_cell_float(wb, sheet_name,row,col):
            return float(get_cell_str(wb, sheet_name,row,col).replace(' ','').replace(',','.')) if get_cell_str(wb, sheet_name,row,col).replace(' ','').replace(',','.') else float(0.0)

        sheet_names = get_sheets(cursor)

        wb = load_workbook(file_path)

        try:
            wb.save(file_path)
        except Exception:
            conn.close()
            root.after(0, loading.destroy)
            root.after(0, lambda: main_menu(root, name_db))
            messagebox.showinfo("Помилка", "Файл відкрито в іншій програмі!")
            return 

        people = []
        vacancies = []

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                wb.save(file_path)


            for i in range(2, wb[sheet_name].max_row+1):
                value = str(get_cell_str(wb, sheet_name, i, 1)).strip().replace('  ',' ')

                if not value:
                    continue

                value1 = ""
                value2 = ""
                value3 = ""

                j=0
 
                while j < len(value) and value[j] != ' ' :
                    value1 += value[j]
                    j += 1
                j +=1
                while j < len(value) and value[j] != ' ' :
                    value2 += value[j]
                    j += 1
                j +=1
                while j < len(value) and value[j] != ' ' :
                    value3 += value[j]
                    j += 1
                    
                if not value3:
                    if value1.lower() != 'вакансія':
                        if value1.lower()[:8] != 'вакансія':
                            continue
                        value2 = value1[8:]
                    value1 = "Вакансія"
                    try:
                        number = int(value2)
                    except Exception:
                        continue

                department = str(get_cell_str(wb, sheet_name, i, 3)).replace(' ','')

                if not department:
                    continue

               
                position = str(get_cell_str(wb, sheet_name, i, 2)).replace(' ','')

                if position[0].lower() == 'з':
                    position = "заф. кафедри"
                elif position[0].lower() == 'п':
                    position = "професор"
                elif position[0].lower() == 'д':
                    position = "доцент"
                elif position[0].lower() == 'с':
                    position = "ст. викладач"
                elif position[0].lower() == 'а':
                    position = "асистент"
                elif position[0].lower() == 'в':
                    if position[1].lower()  == '.':
                        position = "в.о. заф. кафедри"
                    elif position[1].lower() == 'и':
                        position = "викладач"
                    else:
                        continue
                else:
                    continue


                if value3:
                    cursor.execute("""
                        SELECT academic_year.'ставка', academic_year.'всього' , academic_year.'job_id', job.'НПП_ПП'

                        FROM academic_year
                        
                        JOIN job ON academic_year."job_id" = job.id
                        JOIN code_year ON job."код_рік_id" = code_year.id
                        JOIN person ON job.id = person."job_id"

                        WHERE person."прізвище" = ? AND person."ім_я" = ? AND person."по-батькові" = ? AND job."посада" = ? AND code_year."назва_кафедри" = ? AND job."знак_сумісності" = ''
                        LIMIT 1
                    """,(value1,value2,value3,position,department))
                    
        
                    result = cursor.fetchall()
                    
                   
                    if not result:
                        continue

                    npp_pp = result[0][3]

                    temp = Person(value1, value2, value3,position,department, "",str(result[0][0]),str(result[0][1]),int(str(result[0][2])),str(result[0][3]))
                   
                    if temp in people:
                        for person in people:
                            if temp==person:
                                cursor.execute("""SELECT load.value FROM load WHERE load.name = ? AND load.npp_pp = ? LIMIT 1""",(sheet_name,str(result[0][3])))
                                result = cursor.fetchall()
                                if not result:
                                    messagebox.showinfo("Помилка", f"Ви намагаєтеся зменшити навантаження для {temp.get_surname()} {temp.get_name()} {temp.get_patronomic()} ({temp.get_department()}), ({npp_pp}).\n\nНа листі: \"{sheet_name}\".\n\nЦей елемент є в базі, але виникає інша помилка. Скоріш за все ви призначили лист \"{sheet_name}\" лише для {"НПП" if npp_pp == "ПП" else "ПП"}, А цей елемент є {npp_pp}.")
                                    continue
                                person.add_less_time(result[0][0])
                                
                                        
                    else:
                       
                        cursor.execute("""SELECT load.value FROM load WHERE load.name = ? AND load.npp_pp = ? LIMIT 1""",(sheet_name,str(result[0][3])))
                        result = cursor.fetchall()
                        if not result:
                            messagebox.showinfo("Помилка", f"Ви намагаєтеся зменшити навантаження для {temp.get_surname()} {temp.get_name()} {temp.get_patronomic()} ({temp.get_department()}), ({npp_pp}).\n\nНа листі: \"{sheet_name}\".\n\nЦей елемент є в базі, але виникає інша помилка. Скоріш за все ви призначили лист \"{sheet_name}\" лише для {"НПП" if npp_pp == "ПП" else "ПП"}, А цей елемент є {npp_pp}.")
                            continue
                        people.append(temp)
                        people[len(people)-1].add_less_time(result[0][0])

                else:
                    cursor.execute("""
                        SELECT academic_year.'ставка', academic_year.'всього' , academic_year.'job_id', job.'НПП_ПП'

                        FROM academic_year
                        
                        JOIN job ON academic_year."job_id" = job.id
                        JOIN code_year ON job."код_рік_id" = code_year.id
                        JOIN vacancy ON job.id = vacancy."job_id"

                        WHERE vacancy."номер" = ? AND job."посада" = ? AND code_year."назва_кафедри" = ? AND job."знак_сумісності" = ''
                        LIMIT 1
                    """,(value2,position,department))
                    
        
                    result = cursor.fetchall()
                    
                   
                    if not result:
                        continue

                    npp_pp = result[0][3]

                    temp = Vacancy(value2,position,department, "",str(result[0][0]),str(result[0][1]),int(str(result[0][2])),str(result[0][3]))

                    if temp in vacancies:
                        for vacancy in vacancies:
                            if temp==vacancy:
                                cursor.execute("""SELECT load.value FROM load WHERE load.name = ? AND load.npp_pp = ? LIMIT 1""",(sheet_name,str(result[0][3])))
                                result = cursor.fetchall()
                                if not result:
                                    messagebox.showinfo("Помилка", f"Ви намагаєтеся зменшити навантаження для Вакансії {temp.get_number()} ({temp.get_department()}), ({npp_pp}).\n\nНа листі: \"{sheet_name}\".\n\nЦей елемент є в базі, але виникає інша помилка. Скоріш за все ви призначили лист \"{sheet_name}\" лише для {"НПП" if npp_pp == "ПП" else "ПП"}, А цей елемент є {npp_pp}.")
                                    continue
                                vacancy.add_less_time(result[0][0])
                    else:
                        
                        
                        cursor.execute("""SELECT load.value FROM load WHERE load.name = ? AND load.npp_pp = ? LIMIT 1""",(sheet_name,str(result[0][3])))
                        result = cursor.fetchall()
                        if not result:
                            messagebox.showinfo("Помилка", f"Ви намагаєтеся зменшити навантаження для Вакансії {temp.get_number()} ({temp.get_department()}), ({npp_pp}).\n\nНа листі: \"{sheet_name}\".\n\nЦей елемент є в базі, але виникає інша помилка. Скоріш за все ви призначили лист \"{sheet_name}\" лише для {"НПП" if npp_pp == "ПП" else "ПП"}, А цей елемент є {npp_pp}.")
                            continue
                        vacancies.append(temp)
                        vacancies[len(vacancies)-1].add_less_time(result[0][0])
                   
     
        cursor.execute("""
            SELECT DISTINCT person.'прізвище', person.'ім_я', person.'по-батькові', job.'посада', code_year.'назва_кафедри', job.'знак_сумісності', academic_year.'ставка', academic_year.'всього', job.id
                
            FROM job

            JOIN code_year ON job."код_рік_id" = code_year.id
            JOIN academic_year ON job.id = academic_year."job_id"
            JOIN person ON job.id = person."job_id"

            ORDER BY code_year.'назва_кафедри'
        """)

        result = cursor.fetchall()

        all_people = []

        for i, value in enumerate(result):
            all_people.append(Person(result[i][0],result[i][1],result[i][2],result[i][3],result[i][4],result[i][5],result[i][6],result[i][7],result[i][8]))
           
        cursor.execute("""
            SELECT DISTINCT vacancy.'номер', job.'посада', code_year.'назва_кафедри', job.'знак_сумісності', academic_year.'ставка',academic_year.'всього', job.id
                
            FROM job

            JOIN code_year ON job."код_рік_id" = code_year.id
            JOIN academic_year ON job.id = academic_year."job_id"
            JOIN vacancy ON job.id = vacancy."job_id"

            ORDER BY code_year.'назва_кафедри'
        """)

        result = cursor.fetchall()
        
        all_vacancies = []

        for i, value in enumerate(result):
            all_vacancies.append(Vacancy(result[i][0],result[i][1],result[i][2],result[i][3],result[i][4],result[i][5],result[i][6]))
       
        all_people = [value for value in all_people if value not in people]
        all_vacancies = [value for value in all_vacancies if value not in vacancies]



        sheet_names = ["Люди зменшення","Вакансії зменшення","Люди без зменшення","Вакансії без зменшення"]

        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
               
            wb.create_sheet(sheet_name)
            wb.save(file_path)
    
        def array_to_do(wb,array, sheet_name,less,file_path):

            if array:
                wb[sheet_name].cell(row=1,column=1,value="ПІБ" if isinstance(array[0], Person) else "Вакансії")

            wb[sheet_name].cell(row=1,column=2,value="Річна ставка")
            wb[sheet_name].cell(row=1,column=3,value="Річні години")
            wb[sheet_name].cell(row=1,column=4,value="НПП/ПП")
            if less == True:
                wb[sheet_name].cell(row=1,column=5,value="Зменшення (год)")
                wb[sheet_name].cell(row=1,column=6,value="Мінімум (год)")
                wb[sheet_name].cell(row=1,column=7,value="Максимум (год)")
                wb[sheet_name].cell(row=1,column=8,value="Загальний мінімум (год)")
                wb[sheet_name].cell(row=1,column=9,value="Загальний максимум (год)")
            else:
                wb[sheet_name].cell(row=1,column=5,value="Загальний мінімум (год)")
                wb[sheet_name].cell(row=1,column=6,value="Загальний максимум (год)")
                

            wb.save(file_path)

            i=2
            for value in array:
                cursor.execute("DELETE FROM check_db WHERE check_db.job_id = ?",(value.get_job_id(),))
                conn.commit()

                cursor.execute("SELECT job.'НПП_ПП' FROM job WHERE job.id = ?",(value.get_job_id(),))

                npp_pp = str(cursor.fetchall()[0][0])

                cursor.execute("SELECT value FROM load WHERE load.name = 'Максимум' AND load.npp_pp = ?",(npp_pp,))

                maximum = float(str(cursor.fetchall()[0][0]).lower().replace(' ','').replace(',','.'))

                cursor.execute("SELECT value FROM load WHERE load.name = 'Мінімум' AND load.npp_pp = ?",(npp_pp,))

                minimum = float(str(cursor.fetchall()[0][0]).lower().replace(' ','').replace(',','.'))

                cursor.execute("SELECT value FROM load WHERE load.name = 'Середнє навантаження' AND load.npp_pp = ?",(npp_pp,))

                midle = float(str(cursor.fetchall()[0][0]).lower().replace(' ','').replace(',','.'))

                cursor.execute("SELECT value FROM load WHERE load.name = 'Максимум перевищення у %' AND load.npp_pp = ?",(npp_pp,))

                max_prosent = float(str(cursor.fetchall()[0][0]).lower().replace(' ','').replace(',','.'))


                if less == True:
                    result = round(midle*value.get_bid()-value.get_less_time())

                    result_min = round(value.get_bid()*minimum)
                    result_max = round(result_min*(100+max_prosent))

                    if result < result_min:
                        result = result_min

                    if result > result_max:
                        result = result_max

  
                result_min_G = round(value.get_bid()*minimum)
                result_max_G = round(value.get_bid()*maximum)    

               
                if less == True:
                    if value.get_total_year() >= result_min and value.get_total_year() <= result_max and value.get_total_year() >= result_min_G and value.get_total_year() <= result_max_G:
                        continue

                    if isinstance(value, Person):
                        temp = f"Зменшення: {result}.\nМінімальна кількість годин: {result_min}.\nМаксимальна кількість годин: {result_max}.\nЗагальний мінімум годин: {result_min_G}.\nЗагальний максимум годин: {result_max_G}.\nДопустимо для {value.get_surname()} {value.get_name()} {value.get_patronomic()}\n\t{value.get_npp_pp()}\n\t{"Не є сумісник" if not value.get_sign() else "Сумісник"}\n\t{value.get_department()}\n\tРічна ставка: {value.get_bid()}\n\tРічні години: {value.get_total_year()}\n?"
                    else:
                        temp = f"Зменшення: {result}.\nМінімальна кількість годин: {result_min}.\nМаксимальна кількість годин: {result_max}.\nЗагальний мінімум годин: {result_min_G}.\nЗагальний максимум годин: {result_max_G}.\nДопустимо для Вакансія {value.get_number()}\n\t{value.get_npp_pp()}\n\е{"Не є сумісник" if not value.get_sign() else "Сумісник"}\n\t{value.get_department()}\n\tРічна ставка: {value.get_bid()}\n\tРічні години: {value.get_total_year()}\n?"
                      
                      
                    if messagebox.askyesno("Попередження",temp):
                        continue
                   
                else:
                    if value.get_total_year() >= result_min_G and value.get_total_year() <= result_max_G:
                        continue


                    if isinstance(value, Person):
                        temp = f"Загальний мінімум годин: {result_min_G}.\nЗагальний максимум годин: {result_max_G}.\nДопустимо для {value.get_surname()} {value.get_name()} {value.get_patronomic()}\n\t{value.get_npp_pp()}\n\t{"Не є сумісник" if not value.get_sign() else "Сумісник"}\n\t{value.get_department()}\n\tРічна ставка: {value.get_bid()}\n\tРічні години: {value.get_total_year()}\n?"
                    else:
                        temp = f"Загальний мінімум годин: {result_min_G}.\nЗагальний максимум годин: {result_max_G}.\nДопустимо для Вакансія {value.get_number()}\n\t{value.get_npp_pp()}\n\t{"Не є сумісник" if not value.get_sign() else "Сумісник"}\n\t{value.get_department()}\n\tРічна ставка: {value.get_bid()}\n\tРічні години: {value.get_total_year()}\n?"
                        
                      
                    if messagebox.askyesno("Попередження",temp):
                        continue

                
                wb[sheet_name].cell(row=i,column=1,value=f"{value.get_surname()} {value.get_name()} {value.get_patronomic()}" if isinstance(array[0], Person) else f"Вакансія {value.get_number()}")
                wb[sheet_name].cell(row=i,column=2,value=f"{value.get_bid()}")
                wb[sheet_name].cell(row=i,column=3,value=f"{value.get_total_year()}")
                wb[sheet_name].cell(row=i,column=4,value=f"{value.get_npp_pp()}")
                if less == True:
                    wb[sheet_name].cell(row=i,column=5,value=f"{result}")
                    wb[sheet_name].cell(row=i,column=6,value=f"{result_min}")
                    wb[sheet_name].cell(row=i,column=7,value=f"{result_max}")
                    wb[sheet_name].cell(row=i,column=8,value=f"{result_min_G}")
                    wb[sheet_name].cell(row=i,column=9,value=f"{result_max_G}")
                else:
                    wb[sheet_name].cell(row=i,column=5,value=f"{result_min_G}")
                    wb[sheet_name].cell(row=i,column=6,value=f"{result_max_G}")

                i += 1
                  
                wb.save(file_path)

                if less == True:
                    cursor.execute("""
                        INSERT INTO check_db ("зменшення", "мінімум", "максимум", "загальний_мінімум", "загальний_максимум", "job_id")
                        VALUES (?, ?, ?, ?, ?, ?)
                        """, (result, result_min, result_max, result_min_G, result_max_G, value.get_job_id()))
                else:
                    cursor.execute("""
                        INSERT INTO check_db ("загальний_мінімум", "загальний_максимум", "job_id")
                        VALUES (?, ?, ?)
                        """, (result_min_G, result_max_G, value.get_job_id()))
                    

                conn.commit()



       
        array_to_do(wb,people,sheet_names[0],True,file_path)
        array_to_do(wb,vacancies,sheet_names[1],True,file_path)

        array_to_do(wb,all_people,sheet_names[2],False,file_path)
        array_to_do(wb,all_vacancies,sheet_names[3],False,file_path)
        
        
        cursor.execute("""
            DELETE 
            FROM check_assistant_teacher
        """)
        conn.commit()
        
        cursor.execute("""
            SELECT
                person.'прізвище',
                person.'ім_я',
                person.'по-батькові',
                job.'посада', 
                job.'знак_сумісності',
                code_year.'назва_кафедри',
                job.'НПП_ПП',
                job.id,
                academic_year.'лекції',
                academic_year.'кваліфікаційна_робота',
                academic_year.'екзамени',
                academic_year.'консультації_перед_екзаменами'
                
            FROM job

            JOIN code_year ON job.'код_рік_id' = code_year.id
            JOIN academic_year ON job.id = academic_year.'job_id'
            JOIN person ON job.id = person.'job_id'

            WHERE 
                (
                    (
                        job.'посада' == 'асистент' AND 
                        ( academic_year.'лекції' > 0 OR academic_year.'кваліфікаційна_робота' > 0 OR academic_year.'екзамени' > 0 OR academic_year.'консультації_перед_екзаменами' > 0 )
                    )

                    OR 

                    (
                        job.'посада' == 'ст. викладач' AND 
                        ( academic_year.'лекції' > 0 OR academic_year.'кваліфікаційна_робота' > 0 )
                    )
                )

            UNION ALL

            SELECT
                vacancy.'назва',
                vacancy.'номер',
                job.'посада', 
                job.'знак_сумісності',
                code_year.'назва_кафедри',
                job.'НПП_ПП',
                job.id,
                academic_year.'лекції',
                academic_year.'кваліфікаційна_робота',
                academic_year.'екзамени',
                academic_year.'консультації_перед_екзаменами',
                -1
      
            FROM job

            JOIN code_year ON job.'код_рік_id' = code_year.id
            JOIN academic_year ON job.id = academic_year.'job_id'
            JOIN vacancy ON job.id = vacancy.'job_id'

            WHERE 
                (
                    (
                        job.'посада' == 'асистент' AND 
                        ( academic_year.'лекції' > 0 OR academic_year.'кваліфікаційна_робота' > 0 OR academic_year.'екзамени' > 0 OR academic_year.'консультації_перед_екзаменами' > 0 )
                    )

                    OR 

                    (
                        job.'посада' == 'ст. викладач' AND 
                        (academic_year.'лекції' > 0 OR academic_year.'кваліфікаційна_робота' > 0)
                    )
                )

            ORDER BY code_year.'назва_кафедри', job.'посада', job.'знак_сумісності'
        """)

        results = cursor.fetchall()

        if not results:
            return

     

        sheet_names = ["Людські помилки в базі","Вакансійні помилки в базі"]

        for sheet_name in sheet_names:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
               
            wb.create_sheet(sheet_name)
            wb[sheet_name].cell(row=1,column=1,value="ПІБ" if sheet_names[0] == sheet_name else "Вакансії")
            wb[sheet_name].cell(row=1,column=2,value="Посада")
            wb[sheet_name].cell(row=1,column=3,value="Знак сумісника")
            wb[sheet_name].cell(row=1,column=4,value="Кафедра")
            wb[sheet_name].cell(row=1,column=5,value="НПП/ПП")
            wb[sheet_name].cell(row=1,column=6,value="Помилка")
            wb.save(file_path)
    

        def error_in_db(wb,array,file_path,sheet_names):
            if not array:
                return
            
            
            i=1
            j=1

            for value in array:
                if value[len(value)-1]!=-1:

                    if str(value[3]).lower().replace(' ','') == 'ст.викладач' and float(str(value[8]).replace(' ','').replace(',','.').lower()) > 0:

                        if messagebox.askyesno("Попередження", f"\t{value[0]} {value[1]} {value[2]} {value[3]}\n\t{"сумісник" if {value[4]} else "не сумісник"}\n\t{value[5]}\n\t{value[6]}\nЧитає лекції, чи є дозвіл?"):
                            continue

                    i+=1
                    sheet_name = sheet_names[0]


                    temp = ""

                    if float(str(value[8]).replace(' ','').replace(',','.').lower()) > 0:
                        temp += f"Читає лекції: {float(str(value[8]).replace(' ','').replace(',','.').lower())} (годин).\n"
                    if float(str(value[9]).replace(' ','').replace(',','.').lower()) > 0:
                        temp += f"Керує кваліфікаційною роботою: {float(str(value[9]).replace(' ','').replace(',','.').lower())} (годин).\n"
                    if str(value[3]).lower().replace(' ','') == 'асистент':
                        if float(str(value[10]).replace(' ','').replace(',','.').lower()) > 0:
                            temp += f"Приймає екзамени: {float(str(value[10]).replace(' ','').replace(',','.').lower())} (годин).\n"
                        if float(str(value[11]).replace(' ','').replace(',','.').lower()) > 0:
                            temp += f"Проводить консультації перед екзаменами: {float(str(value[11]).replace(' ','').replace(',','.').lower())} (годин).\n"
                            

                    wb[sheet_name].cell(row=i,column=1,value=f"{value[0]} {value[1]} {value[2]}")
                    wb[sheet_name].cell(row=i,column=2,value=f"{value[3]}")
                    wb[sheet_name].cell(row=i,column=3,value=f"{value[4]}")
                    wb[sheet_name].cell(row=i,column=4,value=f"{value[5]}")
                    wb[sheet_name].cell(row=i,column=5,value=f"{value[6]}")
                    wb[sheet_name].cell(row=i,column=6,value=f"{temp}")
                   
                    wb.save(file_path)

                    temp = temp.replace('\n',' ').replace('  ',' ')
                    cursor.execute("INSERT INTO check_assistant_teacher('помилка', 'job_id') VALUES (?, ?)""", (str(temp), int(str(value[7]).lower().replace(' ',''))))
                    conn.commit()
                else:
                   
                    if str(value[2]).lower().replace(' ','') == 'ст.викладач' and float(str(value[7]).replace(' ','').replace(',','.').lower()) > 0:
                        
                        if messagebox.askyesno("Попередження", f"\t{value[0]} {value[1]} {value[2]}\n\t{"сумісник" if {value[3]} else "не сумісник"}\n\t{value[4]}\n\t{value[5]}\nЧитає лекції, чи є дозвіл?"):
                            continue

                    j+=1
                    sheet_name = sheet_names[1]

                    temp = ""

                    if float(str(value[7]).replace(' ','').replace(',','.').lower()) > 0:
                        temp += f"Читає лекції: {float(str(value[7]).replace(' ','').replace(',','.').lower())} (годин).\n"
                    if float(str(value[8]).replace(' ','').replace(',','.').lower()) > 0:
                        temp += f"Керує кваліфікаційною роботою: {float(str(value[8]).replace(' ','').replace(',','.').lower())} (годин).\n"
                    if str(value[2]).lower().replace(' ','') == 'асистент':
                        if float(str(value[9]).replace(' ','').replace(',','.').lower()) > 0:
                            temp += f"Приймає екзамени: {float(str(value[9]).replace(' ','').replace(',','.').lower())} (годин).\n"
                        if float(str(value[10]).replace(' ','').replace(',','.').lower()) > 0:
                            temp += f"Проводить консультації перед екзаменами: {float(str(value[10]).replace(' ','').replace(',','.').lower())} (годин).\n"
                            

                    wb[sheet_name].cell(row=j,column=1,value=f"{value[0]} {value[1]}")
                    wb[sheet_name].cell(row=j,column=2,value=f"{value[2]}")
                    wb[sheet_name].cell(row=j,column=3,value=f"{value[3]}")
                    wb[sheet_name].cell(row=j,column=4,value=f"{value[4]}")
                    wb[sheet_name].cell(row=j,column=5,value=f"{value[5]}")
                    wb[sheet_name].cell(row=j,column=6,value=f"{temp}")
                    

                    wb.save(file_path)
                    temp = temp.replace('\n',' ').replace('  ',' ')
                    cursor.execute("INSERT INTO check_assistant_teacher('помилка', 'job_id') VALUES (?, ?)""", (str(temp), int(str(value[6]).lower().replace(' ',''))))
                    conn.commit()
                    
                
                
       
        error_in_db(wb,results,file_path,sheet_names)

        conn.close()
        root.after(0, loading.destroy)
        root.after(0, lambda: main_menu(root, name_db))

    
    from src.settings.settings import loading_window
    from src.menus.main import main_menu

    file_path = filedialog.askopenfilename(
        title="Оберіть Excel файл",
        filetypes=[("Excel File", "*.xlsx")]
    )
    
    if not file_path:
        messagebox.showinfo("Помилка", "Файл не обрано або має валідну помилку в назві!")
        return

    threading.Thread(target=check_file, args=(root, name_db, file_path), daemon=True).start()
   
